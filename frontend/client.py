#Python GUI imports
import tkinter as tk
from tkinter import font, messagebox
#Excel and numbers usage
import pandas as pd
from openpyxl import load_workbook, Workbook
#Screen resolution imports
import pyautogui as pyg
#Graph usage imports
import matplotlib as mpl
mpl.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
#Path checker
from pathlib import Path
#To open the excel file
import os
#Database
from dotenv import load_dotenv
from supabase import create_client, client

#Loads the data from the .env file
load_dotenv()

#Database settings
url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")
Client = create_client(url, key)
#Class containing client code
class StudyAppClient:
    def __init__(self):
        #List with open windows
        self.openWindows = []
        #List with all open popups
        self.openPopups = []
        #Sets main menu window dimensions
        self.windowHeight = 600
        self.windowWidth = 400
        #Sets main menu position on screen relative to screen resolution
        self.screenResolution = pyg.size()
        self.screenMiddleX = int(self.screenResolution[0] / 2 - self.windowWidth / 2)
        self.screenMiddleY = int(self.screenResolution[1] / 2 - self.windowHeight / 2)
        #Excel sheet
        self.dataStorageWorkbookName = "StudyAppDataStorage.xlsx"
        #Default Sheets List
        self.dataStorageInitSheets = ["DASHBOARD DATA", "CATEGORIES"]
        #Default headers for the sheets
        self.dataStorageInitSheetsHeaders = {self.dataStorageInitSheets[0] : [], self.dataStorageInitSheets[1] : ["Category"]}
        #Checks if the excel file already exist
        self.dataStorageWorkbookPath = Path.cwd()/self.dataStorageWorkbookName
        if self.dataStorageWorkbookPath.is_file():
            self.dataStorageWorkbook = load_workbook(self.dataStorageWorkbookPath)
            self.dataStorageWorkbook.save(self.dataStorageWorkbookName)
            print(f"Excel file '{self.dataStorageWorkbookName}' already exists, new one was not made...")
        else:
            #If not creates a new one with the default sheets added
            self.dataStorageWorkbook = Workbook()
            # Sets the name for the file
            self.dataStorageWorkbook.save(self.dataStorageWorkbookName)
            self.dataStorageWorkbook = load_workbook(self.dataStorageWorkbookPath)
            for sheetName in self.dataStorageInitSheets:
                dataSheet = self.dataStorageWorkbook.create_sheet(title=sheetName)
                dataSheet.append(self.dataStorageInitSheetsHeaders[sheetName])
            del self.dataStorageWorkbook["Sheet"]
            self.dataStorageWorkbook.save(self.dataStorageWorkbookName)
            print(f"Excel file '{self.dataStorageWorkbookName}' created!")
        #Init all functions that need to be called when the program is run
        self.ExistingCategories()
        self.UserSignWindow()
    #Function that closes all open windows
    def WindowCloser(self):
        for window in self.openWindows:
            try:
                window.destroy()
            except Exception as e:
                print(f"Error has occurred:\n\n{e}")
    #Function that closes all open popups
    def PopupCloser(self):
        for popup in self.openPopups:
            try:
                popup.destroy()
            except Exception as e:
                print(f"Error has occured:\n\n{e}")
    #Finds the existing categories and appends them to a global reference list
    def ExistingCategories(self):
        self.existingCategories_df = pd.read_excel(self.dataStorageWorkbookName, "CATEGORIES")
        self.categoryList = self.existingCategories_df["Category"].tolist()
    #Function to create a new window
    def NewWindow(self, windowName):
        # Destroys old window
        self.WindowCloser()
        #Clears the open windows list and withdraws the main menu page
        self.openWindows.clear()
        #Spawns new window and sets the position on screen, title, and makes it resizable
        self.newWindow = tk.Toplevel(self.startingWindow)
        self.newWindow.geometry("+%d+%d" % (self.screenMiddleX,self.screenMiddleY))
        self.newWindow.title(windowName)
        self.newWindow.resizable(True, True)
    #Function to create a new popup window
    def NewPopupWindow(self, windowName):
        #Closes all existing popup windows (if there are any, which there shouldn't)
        self.PopupCloser()
        # Spawns new window and sets the position on screen, title, and makes it resizable
        self.newPopupWindow = tk.Toplevel()
        self.newPopupWindow.geometry("+%d+%d" % (self.screenMiddleX, self.screenMiddleY))
        self.newPopupWindow.title(windowName)
        self.newPopupWindow.resizable(False, False)
    #Function that sets the global fonts
    def FontMaker(self):
        self.titleFont = font.Font(family="Poppins", size=36, weight="bold")
        self.buttonFont = font.Font(family="Poppins", size=18, weight="bold")
    #Function to generate the sign-in/sign-up window
    def UserSignWindow(self):
        #Config of the window
        self.startingWindow = tk.Tk()
        #Vars to store username and password
        self.userName = tk.StringVar()
        self.userPass = tk.StringVar()
        self.email = tk.StringVar()
        #Window configs
        self.startingWindow.title("Sign in or Sign up")
        self.startingWindow.geometry(f"400x200+{self.screenMiddleX}+{self.screenMiddleY}")
        self.startingWindow.resizable(False, False)
        #SignUp Function
        def SignUp():
            if self.userName == "" or self.userPass == "" or self.email == "":
                messagebox.showerror("Invalid entry", "One or more fields are incomplete. Unable to sign up")
            elif 
        #Init fonts
        self.FontMaker()
        # Create a frame to center all widgets
        formFrame = tk.Frame(self.startingWindow)
        formFrame.grid(row=0, column=0, padx=20, pady=(10, 10))  # smaller top/bottom margins
        # Center the frame in the window
        self.startingWindow.grid_rowconfigure(0, weight=1)
        self.startingWindow.grid_columnconfigure(0, weight=1)
        # Define a larger bold font
        labelFont = font.Font(family="Poppins", size=11, weight="bold")
        # Email Entry
        emailLabel = tk.Label(formFrame, text="Email", font=labelFont)
        emailLabel.grid(column=0, row=0, pady=(10, 5), padx=10, sticky="e")
        emailEntry = tk.Entry(formFrame, textvariable=self.email, width=30)
        emailEntry.grid(column=1, row=0, pady=(10, 5), padx=10)
        # Username Entry
        userNameEntryLabel = tk.Label(formFrame, text="Username", font=labelFont)
        userNameEntryLabel.grid(column=0, row=1, pady=(5, 5), padx=10, sticky="e")
        userNameEntry = tk.Entry(formFrame, textvariable=self.userName, width=30)
        userNameEntry.grid(column=1, row=1, pady=(5, 5), padx=10)
        # Password Entry
        userNamePassLabel = tk.Label(formFrame, text="Password", font=labelFont)
        userNamePassLabel.grid(column=0, row=2, pady=(5, 5), padx=10, sticky="e")
        userPassEntry = tk.Entry(formFrame, textvariable=self.userPass, show='*', width=30)
        userPassEntry.grid(column=1, row=2, pady=(5, 10), padx=10)
        # Sign Up and Quit Buttons
        buttonFrame = tk.Frame(formFrame)
        buttonFrame.grid(column=0, row=3, columnspan=2, pady=(10, 5))
        signUpButton = tk.Button(buttonFrame, border=5, text="Sign Up", width=12, command=self.SignUp)
        signUpButton.pack(side="right", padx=10)
        quitButton = tk.Button(buttonFrame, border=5, text="Quit", width=12, command=self.ExitProgram)
        quitButton.pack(side="left", padx=10)
        #Inits the starting window
        self.startingWindow.mainloop()
    #Function to create the main menu window
    def MainMenuWindow(self):
        #Opens the mainmenu window
        self.NewWindow("StudyApp")
        #Renames the var
        self.mainMenu = self.newWindow
        #Appends the window to the open windows list
        self.openWindows.append(self.mainMenu)

        title = tk.Label(self.mainMenu, text="StudyApp", font=self.titleFont)
        title.pack(pady = (30, 20))

        self.dashboard = tk.Button(self.mainMenu, height=1, width=12, text="Dashboard", font=self.buttonFont, border=5, command=lambda: self.DashboardWindow())
        self.dashboard.pack(pady=10)
        self.categories = tk.Button(self.mainMenu, height=1, width=12, text="Categories", font=self.buttonFont, border=5, command=lambda: self.CategoriesWindow())
        self.categories.pack(pady=10)
        self.settingsButton = tk.Button(self.mainMenu, height=1, width=12, text="Settings", font=self.buttonFont, border=5, command=lambda: self.SettingsWindowPopup())
        self.settingsButton.pack(pady=10)
        self.exitButton = tk.Button(self.mainMenu, height=1, width=12, text="Exit", font=self.buttonFont, border=5, command=lambda: self.ExitProgram())
        self.exitButton.pack(pady=10)
    #Function to return and open the main menu window again
    def ReturnToMainMenu(self):
        #Closes all currently open windows
        self.WindowCloser()
        #Clear the currently open windows list
        self.openWindows.clear()
        #Reopens the main menu window
        self.MainMenuWindow()
    #Function to create the dashboard window
    def DashboardWindow(self):
        #Init page
        self.NewWindow("Dashboard")
        #Change var name to name of page and appends page name to open windows list
        self.Dashboard = self.newWindow
        self.openWindows.append(self.Dashboard)
        #Page Title
        pageTitle = tk.Label(self.Dashboard, text="Dashboard", font=self.titleFont)
        pageTitle.pack(pady=(30, 20))
        #Study time and record graph
    #Updates the category list to show all new categories
    def CategoriesWindowListUpdate(self, frame):
        for index in range(len(self.categoryList)):
            categoryButton = tk.Button(frame, text=f"{self.categoryList[index]}")
            categoryButton.grid(sticky="ew", pady=2, row=index, column=0)
    #Categories Window displaying all categories and add button
    def CategoriesWindow(self):
        #Init page
        self.NewWindow("Categories")
        #Change var name to name of page and appends page name to open windows list
        self.Categories = self.newWindow
        self.openWindows.append(self.Categories)
        #Page title
        pageTitle = tk.Label(self.Categories, text="Categories", font=self.titleFont)
        pageTitle.grid(pady=20, row=0, column=0)
        #Categories List Frame
        self.categoriesListFrame = tk.Frame(self.Categories, width=400, height=400)
        self.categoriesListFrame.grid(padx=1, row=1, column=0)
        #Category button (for each one)
        try:
            self.CategoriesWindowListUpdate(self.categoriesListFrame)
        except Exception:
            messagebox.showerror("Loading error", "Categories could not be loaded!")
            self.ReturnToMainMenu()
        #Add Categories button
        addCategoriesButton = tk.Button(self.Categories, height=1, text="Add Category", font=self.buttonFont, border=5, command=lambda: self.AddCategories())
        addCategoriesButton.grid(pady=1, padx=10, row=2, sticky="ew")
        #Back to main menu button
        mainMenuButton = tk.Button(self.Categories, height=1, text="Back to main menu", font=self.buttonFont, border=5, command=lambda: self.ReturnToMainMenu())
        mainMenuButton.grid(pady=(1, 15), padx=10, row=3, sticky="ew")

        self.Categories.protocol("WM_DELETE_WINDOW", self.ExitProgram)
    #Add category function definition to add new categories
    def AddCategories(self):
        try:
            #Gives the popup window a proper name
            self.NewPopupWindow("New Category")
            #Change the var name to the name of the popup and appends it to the list
            self.AddCategoriesPopup = self.newPopupWindow
            self.openPopups.append(self.AddCategoriesPopup)
            self.AddCategoriesPopup.geometry("300x120")
            #Input area for name of the category
            categoryName = tk.Entry(self.AddCategoriesPopup, width=30)
            categoryName.grid(sticky="n", pady=5, padx=10, row=0)
            #Submit and cancel functions
            def Submit():
                newCategory = categoryName.get().strip()
                if newCategory not in self.categoryList:
                    self.categoryList.append(newCategory)
                    self.categorySheet = self.dataStorageWorkbook["CATEGORIES"]
                    # Saves combined data to excel
                    try:
                        self.categorySheet.append([newCategory])
                        self.dataStorageWorkbook.save(self.dataStorageWorkbookName)
                        # Confirmation message
                        print(f"New category '{newCategory}' has been added!")
                    except Exception as e:
                        print(f"File could not be saved:\n\n{e}")
                    # Destroys the popup
                    try:
                        self.PopupCloser()
                        self.CategoriesWindowListUpdate(self.categoriesListFrame)
                    except Exception as e:
                        print(f"Error has occurred:\n\n{e}")
                elif newCategory == "":
                    messagebox.showwarning("Empty Category", "Category name cannot be empty!")
                else:
                    #Destroys the "add category" popup
                    self.PopupCloser()
                    #Creates the new popup
                    messagebox.showwarning("Category already exists", "The category entered was not made as it already exists!")
            def Cancel():
                self.PopupCloser()
            #Submit and cancel button
            submitButton = tk.Button(self.AddCategoriesPopup, text="Submit", height=1, width=5, font=self.buttonFont, border=2, command=lambda: Submit())
            submitButton.grid(sticky="se", column=1, row=1)
            cancelButton = tk.Button(self.AddCategoriesPopup, text="Cancel", height=1, width=5, font=self.buttonFont, border=2, command=lambda: Cancel())
            cancelButton.grid(sticky="se", column=0, row=1, padx=3)
        except Exception as e:
            print(f"Error has occurred:\n\n{e}")
    #Function that opens the excel sheet
    def ExcelWindow(self):
        if self.dataStorageWorkbookPath.is_file():
            #Opens the excel file if it exists
            try:
                os.startfile(self.dataStorageWorkbookName)
            except Exception as e:
                print(f"Excel file could not be opened due to an error:\n\n{e}")
        else:
            #Error message
            print(f"Excel file {self.dataStorageWorkbookName} does not exist...")
    #Function that opens a popup with the settings option and excel button
    def SettingsWindowPopup(self):
        #Creates a new popup window and gives it the appropriate title
        self.NewPopupWindow("Settings")
        #Redefines the popup var to a new var for special use case
        self.settingsPopup = self.newPopupWindow
        self.settingsPopup.geometry("300x300")
        #Makes the button to open the excel file
        self.excelButton = tk.Button(self.settingsPopup, height=1, width=12, text="Open Excel", font=self.buttonFont, border=5, command=lambda: self.ExcelWindow())
        self.excelButton.grid(sticky="ew", row=0, column=0)
    #Function to exit the whole program
    def ExitProgram(self):
        self.startingWindow.destroy()
#Starts running the program
if __name__ == "__main__":
    StudyAppClient()
